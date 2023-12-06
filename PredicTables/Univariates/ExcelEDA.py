import os
import openpyxl
from openpyxl.drawing.image import Image
import polars as pl
import pandas as pd

from typing import Union, Tuple

class ExcelEDA:

    def __init__(self,
                 image_folder: str,
                 eda_df: Union[pl.DataFrame, pl.LazyFrame, pd.DataFrame],
                 feature_selection_df: Union[pl.DataFrame, pl.LazyFrame, pd.DataFrame],
                 feature_col_name:str = "feature_name",
                 wb: openpyxl.Workbook = None,
                 title_cell: str = "B1",
                 title: str = "FEATURENAME - Small Business Exploratory Data Analysis",
                 img_cell: str = "B3",
                 img_size: Tuple[int, int] = (400, 300),
                 stats_start_cell: str = "B5",
                 save_path:str = "./eda.xlsx",
                 run_checks: bool = False,
                 new_file_every_n: int = 50,
                 ) -> None:
        """
        Creates an Excel file with EDA information. Each worksheet contains:
            1. Title with the name of the feature
            2. Image of the univariate plots
            3. Univariate statistics
            4. Feature selection statistics

        The worksheets are sorted by the average of the normalized feature selection
        statistics (eg lasso, random forest, etc.) in the feature selection DataFrame.

        Arguments
        ---------
        image_folder: str
            Path to the folder containing the univariate plots. Each plot should be
            named after the feature it represents.
        eda_df: Union[pl.DataFrame, pl.LazyFrame, pd.DataFrame]
            DataFrame containing the univariate statistics. Will be converted to
            pl.LazyFrame if necessary.
        feature_selection_df: Union[pl.DataFrame, pl.LazyFrame, pd.DataFrame]
            DataFrame containing the feature selection statistics. Will be converted
            to pl.LazyFrame if necessary.
        feature_col_name: str, optional
            Name of the column containing the feature names. Default is "feature_name".
        wb: openpyxl.Workbook, optional
            Workbook to write to. If None, a new workbook will be created.
        title_cell: str, optional
            Cell to insert the title into. Default is "B1".
        title: str, optional
            Title of the worksheet. Default is:
            "FEATURENAME - Small Business Exploratory Data Analysis". Note that
            the string "FEATURENAME" will be replaced with the name of the feature, 
            and can be used to customize the title for each worksheet.
        img_cell: str, optional
            Cell to insert the image into. Default is "B3".
        img_size: Tuple[int, int], optional
            Size of the image to insert. Default is (400, 300).
        stats_start_cell: str, optional
            Cell to start inserting the statistics into. Default is "B5".
        save_path: str, optional
            Path to save the Excel file to. Default is "./eda.xlsx".
        run_checks: bool, optional
            Whether to run checks on the inputs. Default is False.
        new_file_every_n: int, optional
            Max number of features to include in each Excel file. Default is 50.
        """
        if self.run_checks:
            assert os.path.isdir(image_folder), \
                f"image_folder {image_folder} must be a valid directory"
            assert isinstance(eda_df, (pl.DataFrame, pl.LazyFrame, pd.DataFrame)), \
                f"eda_df must be a polars or pandas DataFrame (actual: {type(eda_df)})"
            assert isinstance(feature_selection_df, (pl.DataFrame, pl.LazyFrame, pd.DataFrame)), \
                f"feature_selection_df must be a polars or pandas DataFrame \
(actual: {type(feature_selection_df)})"
            assert isinstance(feature_col_name, str), \
                f"feature_col_name must be a string (actual: {type(feature_col_name)})"
            assert isinstance(wb, (openpyxl.Workbook, type(None))), \
                f"wb must be an openpyxl Workbook (actual: {type(wb)})"
            assert isinstance(title_cell, str), \
                f"title_cell must be a string (actual: {type(title_cell)})"
            assert isinstance(title, str), \
                f"title must be a string (actual: {type(title)})"
            assert isinstance(img_cell, str), \
                f"img_cell must be a string (actual: {type(img_cell)})"
            assert isinstance(img_size, tuple), \
                f"img_size must be a tuple (actual: {type(img_size)})"
            assert isinstance(stats_start_cell, str), \
                f"stats_start_cell must be a string (actual: {type(stats_start_cell)})"
            assert isinstance(save_path, str), \
                f"save_path must be a string (actual: {type(save_path)})"

        self.feature_col_name = feature_col_name
        self.image_folder = image_folder
        self.img_cell = img_cell
        self.title_cell = title_cell
        self.title = title
        self.img_size = img_size
        self.stats_start_cell = stats_start_cell
        self.save_path = save_path
        self.run_checks = run_checks
        self.new_file_every_n = new_file_every_n


        # Convert `eda_df` to pl.LazyFrame if necessary
        if isinstance(eda_df, pd.DataFrame):
            self.eda_df = pl.from_pandas(eda_df)
        elif isinstance(eda_df, pl.DataFrame):
            self.eda_df = eda_df.lazy()
        else:
            self.eda_df = eda_df

        if self.run_checks:
            assert self.feature_col_name in self.eda_df.columns, \
                f"eda_df must contain a column named {self.feature_col_name}.\
Actual columns:\n{self.eda_df.columns}"
            assert isinstance(self.eda_df, pl.LazyFrame), \
                f"eda_df should have been converted to a pl.LazyFrame \
(actual: {type(self.eda_df)})"
        
        # Convert `feature_selection_df` to pl.LazyFrame if necessary
        if isinstance(feature_selection_df, pd.DataFrame):
            self.feature_selection_df = pl.from_pandas(feature_selection_df)
        elif isinstance(feature_selection_df, pl.DataFrame):
            self.feature_selection_df = feature_selection_df.lazy()
        else:
            self.feature_selection_df = feature_selection_df

        if self.run_checks:
            assert self.feature_col_name in self.feature_selection_df.columns, \
                f"feature_selection_df must contain a column named {self.feature_col_name}.\
Actual columns:\n{self.feature_selection_df.columns}"
            assert isinstance(self.feature_selection_df, pl.LazyFrame), \
                f"feature_selection_df should have been converted to a pl.LazyFrame \
(actual: {type(self.feature_selection_df)})"

        # Create Excel workbook
        if wb is None:
            self._create_excel_workbook()
        else:
            self.wb = wb

        if self.run_checks:
            assert self.wb is not None, \
    f"wb should not be None (actual: {type(self.wb)})"
            assert isinstance(self.wb, openpyxl.Workbook), \
                f"wb should be an openpyxl Workbook (actual: {type(self.wb)})"

    def _resize_image(self, img: Image) -> Image:
        """Resizes an image to `self.img_size`."""
        if self.run_checks:
            # Validate inputs
            assert isinstance(img, Image), \
                f"img must be an openpyxl Image (actual: {type(img)})"
            assert isinstance(self.img_size, tuple), \
                f"img_size must be a tuple (actual: {type(self.img_size)})"
            assert len(self.img_size) == 2, \
                f"img_size must be a tuple of length 2 (actual: {len(self.img_size)})"
            assert isinstance(self.img_size[0], int), \
                f"img_size[0] must be an int (actual: {type(self.img_size[0])})"
            assert isinstance(self.img_size[1], int), \
                f"img_size[1] must be an int (actual: {type(self.img_size[1])})"
        
        # Resize image
        img.width, img.height = self.img_size
        
        # Return resized image
        if self.run_checks:
            assert img.width == self.img_size[0], \
                f"img.width should be {self.img_size[0]} (actual: {img.width})"
            assert img.height == self.img_size[1], \
                f"img.height should be {self.img_size[1]} (actual: {img.height})"
            return img        

    def _get_image_files(self) -> list:
        """Returns a list of image files in the image folder."""
        if self.run_checks:
            # Validate inputs
            assert os.path.isdir(self.image_folder), \
                f"image_folder {self.image_folder} must be a valid directory"
            assert len(os.listdir(self.image_folder)) > 0, \
                f"image_folder {self.image_folder} must contain at least one image"
            assert any([f.endswith('.png') for f in os.listdir(self.image_folder)]), \
                f"image_folder {self.image_folder} must contain at least one .png file"
        
        # Return list of image files
        img_list = [f for f in os.listdir(self.image_folder) if f.endswith('.png')]
        if self.run_checks:
            assert len(img_list) > 0, \
                f"image_folder {self.image_folder} must contain at least one .png file"
        return img_list
    
    def _get_image_path(self, feature: str) -> str:
        """Returns the path to the image file for a feature."""
        if self.run_checks:
            # Validate inputs
            assert isinstance(feature, str), \
                f"feature must be a string (actual: {type(feature)})"
            assert feature in self._get_feature_names(), \
                f"feature {feature} not found in image_folder {self.image_folder}\n\n\
Available features:\n{self._get_feature_names()}"
            assert os.path.isfile(os.path.join(self.image_folder, feature + '.png')), \
                f"image file {feature + '.png'} not found in image_folder \
{self.image_folder}. Only .png files are supported."
        
        # Return path to image file
        img_pth = os.path.join(self.image_folder, feature + '.png')

        if self.run_checks:
            assert os.path.isfile(img_pth), \
                f"image file {img_pth} not found in image_folder {self.image_folder}"
        return img_pth
    
    def _get_feature_names(self) -> list:
        """Returns a list of feature names from the image files."""
        if self.run_checks:
            # Validate inputs
            assert len(self._get_image_files()) > 0, \
                f"image_folder {self.image_folder} must contain at least one image"
            assert any([f.endswith('.png') for f in self._get_image_files()]), \
                f"image_folder {self.image_folder} must contain at least one .png file"
        
        # Return list of feature names
        feature_names = [os.path.splitext(f)[0] for f in self._get_image_files()]

        if self.run_checks:
            assert len(feature_names) > 0, \
                f"image_folder {self.image_folder} must contain at least one .png file"
            assert all([isinstance(f, str) for f in feature_names]), \
                f"feature names must be strings (actual: \
{[f'{f}: {type(f)}' for f in feature_names]})"
        return feature_names
    
    def _normalize_single_feature(self, df: pl.LazyFrame, feature:str) -> pl.LazyFrame:
        """Normalizes the feature selection statistics for a single feature."""
        if self.run_checks:
            # Validate inputs
            assert isinstance(df, pl.LazyFrame), \
                f"df must be a polars LazyFrame (actual: {type(df)})"
            assert isinstance(feature, str), \
                f"feature must be a string (actual: {type(feature)})"
            assert feature in self._get_feature_names(), \
                f"feature {feature} not found in image_folder {self.image_folder}\n\n\
Available features:\n{self._get_feature_names()}"
            assert feature in df[self.feature_col_name].to_list(), \
                f"feature {feature} not found in df\n\n\
Available features:\n{df[self.feature_col_name].to_list()}"

        feature_df = df.select([feature])\
                        .with_columns([
                            pl.divide(pl.col(feature) - pl.col(feature).min(),
                                      pl.col(feature).max() - pl.col(feature).min())\
                                .alias("normalized_stat")])
        
        # Return normalized feature selection statistics
        if self.run_checks:
            assert feature_df.shape[0] == df.shape[0], \
                f"feature_df should have the same number of rows as df \
(actual: {feature_df.shape[0]} vs {df.shape[0]})"
            assert feature_df.shape[1] == df.shape[1] + 1, \
                f"feature_df should have one more column than df (representing the \
normalized feature selection statistic)\n(actual: {feature_df.shape[1]} vs \
{df.shape[1] + 1})"
            assert feature_df["normalized_stat"].min() == 0, \
                f"normalized_stat should have a minimum of 0 (actual: \
{feature_df['normalized_stat'].min()})"
            assert feature_df["normalized_stat"].max() == 1, \
                f"normalized_stat should have a maximum of 1 (actual: \
{feature_df['normalized_stat'].max()})"
            assert isinstance(feature_df, pl.LazyFrame), \
                f"feature_df should be a polars LazyFrame (actual: {type(feature_df)})"
        return feature_df
    
    def _normalize_feature_selection_df(self, df: pl.LazyFrame) -> pl.LazyFrame:
        """Normalizes the feature selection statistics for all features."""
        if self.run_checks:
            # Validate inputs
            assert isinstance(df, pl.LazyFrame), \
                f"df must be a polars LazyFrame (actual: {type(df)})"
            assert self.feature_col_name in df.columns, \
                f"df must contain a column named {self.feature_col_name}\
Actual columns:\n{df.columns}"
        
        normalized_df = df.select("feature_name")\
                          .with_columns([
                              self._normalize_single_feature(df, feature) \
                              for feature in self._get_feature_names()
                          ])
        
        # Return normalized feature selection statistics
        if self.run_checks:
            assert normalized_df.shape[0] == df.shape[0], \
                f"normalized_df should have the same number of rows as df \
(actual: {normalized_df.shape[0]} vs {df.shape[0]})"
            assert normalized_df.shape[1] == df.shape[1], \
                f"normalized_df should have the same number of columns as df \
(actual: {normalized_df.shape[1]} vs {df.shape[1]})"
            assert all([f in normalized_df.columns for f in self._get_feature_names()]), \
                f"normalized_df should contain a column for each feature \
(columns with feature names: {normalized_df.columns}\n\
features: {self._get_feature_names()})\n\n\
columns without feature names: \
{[c for c in normalized_df.columns if c not in self._get_feature_names()]}"
            assert all([normalized_df[f].min() == 0 for f in self._get_feature_names()]), \
                f"normalized_df should have a minimum of 0 for each feature \
(actual: {[normalized_df[f].min() for f in self._get_feature_names()]})"
            assert all([normalized_df[f].max() == 1 for f in self._get_feature_names()]), \
                f"normalized_df should have a maximum of 1 for each feature \
(actual: {[normalized_df[f].max() for f in self._get_feature_names()]})"
            assert isinstance(normalized_df, pl.LazyFrame), \
                f"normalized_df should be a polars LazyFrame (actual: {type(normalized_df)})"
        
        return normalized_df
    
    def _average_normalized_stats(self) -> pl.LazyFrame:
        """Averages the normalized feature selection statistics."""
        if self.run_checks:
            # Validate inputs
            assert isinstance(self.feature_selection_df, pl.LazyFrame), \
                f"feature_selection_df should be a polars LazyFrame \
(actual: {type(self.feature_selection_df)})"
            assert self.feature_col_name in self.feature_selection_df.columns, \
                f"feature_selection_df must contain a column named {self.feature_col_name}\
Actual columns:\n{self.feature_selection_df.columns}"
        
        df = self._normalize_feature_selection_df(self.feature_selection_df)
        df["avg_stat"] = df.mean(axis=1)

        # Return DataFrame with average normalized feature selection statistics
        if self.run_checks:
            assert df.shape[0] == self.feature_selection_df.shape[0], \
                f"df should have the same number of rows as feature_selection_df \
(actual: {df.shape[0]} vs {self.feature_selection_df.shape[0]})"
            assert df.shape[1] == self.feature_selection_df.shape[1] + 1, \
                f"df should have one more column than feature_selection_df \
(actual: {df.shape[1]} vs {self.feature_selection_df.shape[1] + 1})"
            assert "avg_stat" in df.columns, \
                f"df should contain a column named 'avg_stat' (actual: {df.columns})"
            assert df["avg_stat"].min() == 0, \
                f"avg_stat should have a minimum of 0 (actual: {df['avg_stat'].min()})"
            assert df["avg_stat"].max() == 1, \
                f"avg_stat should have a maximum of 1 (actual: {df['avg_stat'].max()})"
            assert isinstance(df, pl.LazyFrame), \
                f"df should be a polars LazyFrame (actual: {type(df)})"
        return df
    
    def _sort_features(self) -> list:
        """Sorts the features by the average normalized feature selection statistic."""
        if self.run_checks:
            # Validate inputs
            assert isinstance(self.feature_selection_df, pl.LazyFrame), \
                f"feature_selection_df should be a polars LazyFrame \
(actual: {type(self.feature_selection_df)})"
            assert self.feature_col_name in self.feature_selection_df.columns, \
                f"feature_selection_df must contain a column named {self.feature_col_name}\
Actual columns:\n{self.feature_selection_df.columns}"
        
        df = self._average_normalized_stats()
        sorted_features = df.sort("avg_stat", reverse=True)["feature_name"].to_list()

        # Return list of features sorted by average normalized feature selection
        # statistic
        if self.run_checks:
            assert len(sorted_features) == df.shape[0], \
                f"sorted_features should have the same number of rows as df \
(actual: {len(sorted_features)} vs {df.shape[0]})"
            assert all([f in sorted_features for f in self._get_feature_names()]), \
                f"sorted_features should contain a feature for each feature \
(features in sorted_features: {sorted_features}\n\
features: {self._get_feature_names()})\n\n\
features not in sorted_features: \
{[f for f in self._get_feature_names() if f not in sorted_features]}"
            assert isinstance(sorted_features, list), \
                f"sorted_features should be a list (actual: {type(sorted_features)})"
            return sorted_features
    
    def _create_excel_workbook(self) -> None:
        """Creates the Excel workbook."""
        if self.run_checks:
            # Validate inputs
            assert isinstance(self.wb, (openpyxl.Workbook, type(None))), \
                f"wb must be an openpyxl Workbook (actual: {type(self.wb)})"
        
        if self.wb is None:
            self.wb = openpyxl.Workbook()
            # Remove default sheet
            self.wb.remove(self.wb.active)

        # Validate Excel workbook
        if self.run_checks:
            assert isinstance(self.wb, openpyxl.Workbook), \
                f"wb should be an openpyxl Workbook (actual: {type(self.wb)})"
            assert len(self.wb.sheetnames) == 0, \
                f"wb should not contain any sheets (actual: {len(self.wb.sheetnames)})"
            assert self.wb.active is None, \
                f"wb should not have an active sheet (actual: {self.wb.active})"
    
    def _create_single_worksheet(self, feature: str):
        """Creates a single worksheet for a feature."""
        if self.run_checks:
            # Validate inputs
            assert isinstance(feature, str), \
                f"feature must be a string (actual: {type(feature)})"
            assert feature in self._get_feature_names(), \
                f"feature {feature} not found in image_folder {self.image_folder}\n\n\
Available features:\n{self._get_feature_names()}"
            assert isinstance(self.wb, openpyxl.Workbook), \
                f"wb should be an openpyxl Workbook (actual: {type(self.wb)})"
            assert len(self.wb.sheetnames) == 0, \
                f"wb should not contain any sheets (actual: {len(self.wb.sheetnames)})"
        
        # Create new sheet for feature
        ws = self.wb.create_sheet(title=feature)

        # Add title, replacing "FEATURENAME" with the feature name
        ws[self.title_cell] = self.title.replace("FEATURENAME", feature)

        # Insert image
        img_path = self._get_image_path(feature)
        img = Image(img_path)
        ws.add_image(img, self.img_cell)

        # Resize image
        img = self._resize_image(img)

        # Add univariate stats
        stats = self.eda_df\
                    .filter(self.eda_df["feature_name"] == feature)\
                    .to_dict(orient="records")[0]
        
        for i, (k, v) in enumerate(stats.items()):
            start_cell = f"{self.stats_start_cell[0]}\
{i + self.stats_start_cell[1:]}"
            end_cell = f"{chr(ord(self.stats_start_cell[0]) + 1)}\
{i + self.stats_start_cell[1:]}"
            ws[start_cell] = k
            ws[end_cell] = v

        # Add feature selection stats
        fs_stats = self.feature_selection_df\
                        .filter(self.feature_selection_df["feature_name"] == feature)\
                        .to_dict(orient="records")[0]
        
        for i, (k, v) in enumerate(fs_stats.items(), start=len(stats)):
            start_cell = f"{self.stats_start_cell[0]}\
{i + self.stats_start_cell[1:]}"
            end_cell = f"{chr(ord(self.stats_start_cell[0]) + 1)}\
{i + self.stats_start_cell[1:]}"
            ws[start_cell] = k
            ws[end_cell] = v

    def _create_worksheets(self) -> None:
        """Creates a worksheet for each feature."""
        for feature in self._sort_features():
            self._create_single_worksheet(feature)

    def _save_chunks_to_files(self):
        """Saves the workbook to multiple files."""
        # Chunk the features into groups of size new_file_every_n
        all_features = self._sort_features()
        chunks = [all_features[i:i+self.new_file_every_n] 
                for i in range(0, len(all_features), self.new_file_every_n)]

        # For each chunk of features, create a new Excel workbook, 
        # create worksheets for those features, and save it to a unique file.
        for i, chunk in enumerate(chunks):
            self.wb = openpyxl.Workbook()  # create a new workbook
            self.wb.remove(self.wb.active)  # remove the default worksheet
            
            for feature in chunk:
                self._create_single_worksheet(feature)

            # Modify the save path to have a unique name for each chunk
            # e.g., if save_path is "./eda.xlsx", filenames would be
            # "./eda_1.xlsx", "./eda_2.xlsx", etc.
            base_name, ext = os.path.splitext(self.save_path)
            chunked_save_path = f"{base_name}_{i + 1}{ext}"
            self.wb.save(chunked_save_path)


    # def save(self, path: str) -> None:
    #     """Saves the workbook to `path`."""
    #     self._create_worksheets()
    #     self.wb.save(path)

    def save(self):
        self._save_chunks_to_files()
